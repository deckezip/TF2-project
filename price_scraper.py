# -*- coding: utf-8 -*-

import json
import pandas as pd
import xlwt
import datetime
import requests
import steam.webauth as wa
from openpyxl import Workbook


class PriceScraper(object):
    MARKET_URL = 'https://steamcommunity.com/market/'
    PRICEHISTORY_URL = MARKET_URL + 'pricehistory/?country={}&currency={}&appid={}&market_hash_name={}'
    PRICEOVERVIEW_URL = MARKET_URL + 'priceoverview/??market_hash_name={}&appid={}&currency={}'

    def __init__(self, login, password, app_id=440, currency=1, country='RU'):
        self.APP_ID = app_id
        self.CURRENCY = currency
        self.COUNTRY = country
        self._login = login
        self._password = password
        user = wa.WebAuth(login)
        self._session = user.cli_login(password)

    def get_item_history(self, item_hash_name):
        url = self.PRICEHISTORY_URL.format(self.COUNTRY, self.CURRENCY, self.APP_ID, item_hash_name)
        r = self._session.get(url)
        # print(r.text, r.status_code)
        if r.status_code == 200 or r.status_code == 500:
            # print(type(json.loads(r.text)))
            return True, json.loads(r.text)
        elif r.status_code == 429:
            print('too many requests')
            return False, None

    def get_item_priceoverview(self, item_hash_name):
        url = self.PRICEOVERVIEW_URL.format(item_hash_name, self.APP_ID, self.CURRENCY)

    # get price info in format ["Oct 22 2020 01: +0",0.098,"1"] and convert to (datetime(2020, 10, 22, 0, 0), 0.098, 1)
    @staticmethod
    def convert_price(price):
        day_price = ' '.join(price[0].split(' ')[:3])
        day = datetime.datetime.strptime(day_price, '%b %d %Y')
        return day, price[1], int(price[2])

    def filter_prices(self, prices, year=2017):
        converted_prices = []
        for price in prices:
            new_price = self.convert_price(price)
            if new_price[0].year >= year:
                converted_prices.append(new_price)
        # print(converted_prices)
        filtered_prices = []
        if len(converted_prices):
            day_prices = [converted_prices[0]]
            for i in range(1, len(converted_prices)):
                if converted_prices[i][0] == day_prices[0][0]:
                    day_prices.append(converted_prices[i])
                else:
                    if len(day_prices) > 1:
                        filtered_prices.append(self.normalize_day_price(day_prices))
                    else:
                        filtered_prices.append(day_prices[0])
                    day_prices.clear()
                    day_prices.append(converted_prices[i])

            if len(day_prices):
                filtered_prices.append(day_prices[0])

        return filtered_prices

    @staticmethod
    def generate_days(start_date, end_date):
        date = start_date
        dates = []
        while date <= end_date:
            dates.append(date.strftime("%m-%d-%Y"))
            date = date + datetime.timedelta(days=1)
        return dates

    @staticmethod
    def normalize_prices(prices, start_date, now_date):
        normalized_prices = []
        prev_price_volume = 0.0
        date = start_date
        price_iter = iter(prices)
        cur_price = next(price_iter)
        while date <= now_date:
            if cur_price is not None and date == cur_price[0]:
                normalized_prices.append(cur_price)
                prev_price_volume = cur_price[1]
                try:
                    cur_price = next(price_iter)
                except StopIteration:
                    cur_price = None
            else:
                normalized_prices.append((date, prev_price_volume, 0))

            date = date + datetime.timedelta(days=1)

        return normalized_prices

    @staticmethod
    def normalize_day_price(day_prices):
        # print('day prices', day_prices)
        sales = sum([price[1] * price[2] for price in day_prices])
        count_sales = sum([price[2] for price in day_prices])
        return day_prices[0][0], round(sales / count_sales, 3), count_sales

    def save_first_column(self, work_sheet, column):
        for i, item in enumerate(column):
            work_sheet.cell(i + 2, 1).value = column[i]

    def save_item_history(self, work_sheet, column_index, item_name, data, data_index):
        work_sheet.cell(1, column_index + 1).value = item_name
        for i, item in enumerate(data):
            work_sheet.cell(i + 2, column_index + 1).value = item[data_index]

    def save_price_overview(self):
        pass

    def scrape_history(self, items_names):
        # init date sequence
        now = datetime.datetime.now()
        start_date = datetime.datetime(2017, 1, 1)
        now_date = datetime.datetime(now.year, now.month, now.day)
        dates = self.generate_days(start_date, now_date)

        # create data frames
        df_prices = pd.DataFrame(columns=[' '] + items_names)
        df_sales = pd.DataFrame(columns=[' '] + items_names)

        # init excel Workbooks
        wb_prices = xlwt.Workbook()
        wb_sales = xlwt.Workbook()
        ws_prices = wb_prices.add_sheet('Prices')
        ws_sales = wb_sales.add_sheet('Item sales')

        workbook_prices = Workbook()
        workbook_sales = Workbook()
        worksheet_prices = workbook_prices.active
        worksheet_sales = workbook_sales.active

        # save first column with dates sequence in worksheets
        worksheet_prices.append([' '] + items_names)
        worksheet_sales.append([' '] + items_names)

        # df_prices[' '] = dates
        # df_sales[' '] = dates
        self.save_first_column(worksheet_prices, dates)
        self.save_first_column(worksheet_sales, dates)

        # scrape history and write to file
        count_items = len(items_names)
        name_index = 0
        while name_index < count_items:
            name = items_names[name_index]
            print('request to {}/{} - name: {}'.format(name_index + 1, count_items, name.strip('\n')))
            res, history = self.get_item_history(name)
            if res:
                if history['success']:
                    name_index += 1
                    data = self.normalize_prices(self.filter_prices(history['prices']), start_date, now_date)
                    self.save_item_history(worksheet_prices, name_index, name, data, 1)
                    self.save_item_history(worksheet_sales, name_index, name, data, 2)
                    # df_prices[name] = [item[1] for item in data]
                    # df_sales[name] = [item[2] for item in data]
                    # worksheet_prices.append([name] + [item[1] for item in data])
                    # worksheet_sales.append([name] + [item[2] for item in data])
                    print('Save history of {} '.format(name))

        # with pd.ExcelWriter('prices.xlsx') as writer:
        #     df_prices.to_excel(writer, sheet_name='Prices', index=False)
        # with pd.ExcelWriter('sales.xlsx') as writer:
        #     df_sales.to_excel(writer, sheet_name='Sales', index=False)

        workbook_prices.save('./prices.xlsx')
        workbook_sales.save('./sales.xlsx')

        # save files with data
        # wb_prices.save('prices.xls')
        # wb_sales.save('sales.xls')

login = ''
password = ''


def filter_names():
    scraper = PriceScraper(login, password)
    with open('./market_names.txt', 'r') as f:
        names = f.readlines()

    file = open('./filter_names.txt', 'a')
    try:
        i = 21857
        while i < len(names):
            item_name = names[i]
            print('request to {}/{}  - name: {}'.format(i, len(names), item_name))
            res, data = scraper.get_item_history(item_name)
            # print(res, data)
            if res and data:
                if data['success'] and len(data['prices']) >= 1500:
                    file.write('{}'.format(item_name))
                    print('Save {}'.format(item_name))
                elif data['success']:
                    print('Too few prices {}'.format(len(data['prices'])))
                i += 1
                print('success request')
            else:
                print('bad request')
    except Exception as e:
        print(e)
        file.close()
    finally:
        file.close()


def main():
    scraper = PriceScraper(login, password)
    with open('./filter_names.txt', 'r') as f:
        names = f.readlines()
    scraper.scrape_history(list(set(names)))

    # prices = [["Jun 07 2017 01: +0",0.252,"2"],["Jun 07 2017 01: +0",0.27,"2"],["Jun 09 2017 01: +0",0.28,"1"],["Jun 10 2017 01: +0",0.279,"1"],["Jun 11 2017 01: +0",0.259,"4"]]
    # print(scraper.filter_prices(prices))
    # print(scraper.get_item_history('The Diplomat'))


if __name__ == '__main__':
    # filter_names()
    main()

